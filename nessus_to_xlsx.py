#!/usr/bin/env python3
"""
Nessus XML to XLSX Converter
Konvertiert .nessus Dateien (XML) in Excel XLSX Format

Verwendung:
    python nessus_to_xlsx.py input.nessus
    python nessus_to_xlsx.py --dir /pfad/zu/nessus/dateien
    python nessus_to_xlsx.py input.nessus -o custom_name.xlsx

Abhängigkeiten:
    pip install openpyxl
"""

import argparse
import sys
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class NessusParser:
    """Parser für Nessus XML Dateien"""
    
    SEVERITY_MAP = {
        '0': 'Info',
        '1': 'Low',
        '2': 'Medium',
        '3': 'High',
        '4': 'Critical'
    }
    
    SEVERITY_COLORS = {
        'Critical': 'C00000',  # Rot
        'High': 'FF6600',      # Orange
        'Medium': 'FFCC00',    # Gelb
        'Low': '92D050',       # Grün
        'Info': '00B0F0'       # Blau
    }
    
    def __init__(self, nessus_file: Path):
        self.nessus_file = nessus_file
        self.tree = None
        self.root = None
        self.hosts = []
        self.vulnerabilities = defaultdict(list)
        
    def parse(self) -> bool:
        """Parse die Nessus XML Datei"""
        try:
            print(f"📂 Parse {self.nessus_file.name} ({self.nessus_file.stat().st_size / 1024 / 1024:.1f} MB)...")
            print(f"   Lade XML (kann bei großen Dateien dauern)...")
            
            self.tree = ET.parse(self.nessus_file)
            self.root = self.tree.getroot()
            
            # Prüfe Nessus Version
            if self.root.tag != 'NessusClientData_v2':
                print(f"❌ Keine gültige Nessus v2 Datei!")
                return False
            
            # Parse Report(s) - kann mehrere geben
            reports = self.root.findall('.//Report')
            if not reports:
                print(f"❌ Kein Report Element gefunden!")
                return False
            
            print(f"   Gefunden: {len(reports)} Report(s)")
            
            total_items = 0
            # Parse alle Reports
            for report in reports:
                report_hosts = report.findall('ReportHost')
                print(f"   Verarbeite {len(report_hosts)} Hosts...")
                
                # Parse Hosts
                for idx, report_host in enumerate(report_hosts, 1):
                    if idx % 10 == 0:
                        print(f"   ... Host {idx}/{len(report_hosts)}")
                    
                    host_data = self._parse_host(report_host)
                    
                    # Prüfe ob Host schon existiert (bei kombinierten Scans)
                    existing_host = None
                    for h in self.hosts:
                        if h['name'] == host_data['name'] or h['ip'] == host_data['ip']:
                            existing_host = h
                            break
                    
                    if existing_host is None:
                        self.hosts.append(host_data)
                    
                    # Sammle ALLE Vulnerabilities
                    items = report_host.findall('ReportItem')
                    total_items += len(items)
                    
                    for item in items:
                        vuln_data = self._parse_vulnerability(item, host_data)
                        severity_num = vuln_data['severity']
                        severity = self.SEVERITY_MAP.get(severity_num, 'Info')
                        
                        # Filtere nur echte Vulnerabilities (severity > 0)
                        # ABER: auch 0 (Info) kann wichtig sein
                        self.vulnerabilities[severity].append(vuln_data)
            
            print(f"\n✅ {len(self.hosts)} Hosts gefunden")
            print(f"   Total ReportItems: {total_items}")
            print(f"   Critical (4): {len(self.vulnerabilities['Critical'])}")
            print(f"   High (3): {len(self.vulnerabilities['High'])}")
            print(f"   Medium (2): {len(self.vulnerabilities['Medium'])}")
            print(f"   Low (1): {len(self.vulnerabilities['Low'])}")
            print(f"   Info (0): {len(self.vulnerabilities['Info'])}")
            
            # Debug: Zeige Severity-Verteilung aus den tatsächlichen Items
            severity_counts = {}
            for report in reports:
                for report_host in report.findall('ReportHost'):
                    for item in report_host.findall('ReportItem'):
                        sev = item.get('severity', '0')
                        severity_counts[sev] = severity_counts.get(sev, 0) + 1
            
            if severity_counts:
                print(f"\n   Debug - Raw XML Severity Counts:")
                for sev in ['4', '3', '2', '1', '0']:
                    if sev in severity_counts:
                        print(f"   Severity {sev}: {severity_counts[sev]}")
            
            return True
            
        except ET.ParseError as e:
            print(f"❌ XML Parse Fehler: {e}")
            return False
        except Exception as e:
            print(f"❌ Fehler beim Parsen: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _parse_host(self, report_host) -> Dict[str, Any]:
        """Parse Host Informationen"""
        host_data = {
            'name': report_host.get('name'),
            'ip': '',
            'fqdn': '',
            'os': '',
            'mac': '',
            'netbios': ''
        }
        
        # Parse HostProperties
        properties = report_host.find('HostProperties')
        if properties is not None:
            for tag in properties.findall('tag'):
                name = tag.get('name')
                value = tag.text or ''
                
                if name == 'host-ip':
                    host_data['ip'] = value
                elif name == 'host-fqdn':
                    host_data['fqdn'] = value
                elif name == 'operating-system':
                    host_data['os'] = value
                elif name == 'mac-address':
                    host_data['mac'] = value
                elif name == 'netbios-name':
                    host_data['netbios'] = value
        
        return host_data
    
    def _parse_vulnerability(self, item, host_data) -> Dict[str, Any]:
        """Parse Vulnerability Informationen"""
        vuln_data = {
            'host_name': host_data['name'],
            'host_ip': host_data['ip'],
            'host_fqdn': host_data['fqdn'],
            'plugin_id': item.get('pluginID'),
            'plugin_name': item.get('pluginName'),
            'plugin_family': item.get('pluginFamily'),
            'severity': item.get('severity'),
            'port': item.get('port'),
            'protocol': item.get('protocol'),
            'service': item.get('svc_name'),
            'synopsis': '',
            'description': '',
            'solution': '',
            'cvss': '',
            'cvss_vector': '',
            'cve': '',
            'bid': '',
            'xref': '',
            'see_also': '',
            'plugin_output': ''
        }
        
        # Parse Details
        for child in item:
            tag = child.tag
            value = child.text or ''
            
            if tag == 'synopsis':
                vuln_data['synopsis'] = value
            elif tag == 'description':
                vuln_data['description'] = value
            elif tag == 'solution':
                vuln_data['solution'] = value
            elif tag == 'cvss_base_score':
                vuln_data['cvss'] = value
            elif tag == 'cvss_vector':
                vuln_data['cvss_vector'] = value
            elif tag == 'cve':
                vuln_data['cve'] = vuln_data['cve'] + value + ', ' if vuln_data['cve'] else value
            elif tag == 'bid':
                vuln_data['bid'] = vuln_data['bid'] + value + ', ' if vuln_data['bid'] else value
            elif tag == 'xref':
                vuln_data['xref'] = vuln_data['xref'] + value + ', ' if vuln_data['xref'] else value
            elif tag == 'see_also':
                vuln_data['see_also'] = value
            elif tag == 'plugin_output':
                vuln_data['plugin_output'] = value
        
        # Entferne trailing commas
        vuln_data['cve'] = vuln_data['cve'].rstrip(', ')
        vuln_data['bid'] = vuln_data['bid'].rstrip(', ')
        vuln_data['xref'] = vuln_data['xref'].rstrip(', ')
        
        return vuln_data


class ExcelWriter:
    """Schreibt geparste Daten in Excel"""
    
    def __init__(self, output_file: Path, parser: NessusParser):
        self.output_file = output_file
        self.parser = parser
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Entferne default Sheet
        
        # Styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def write(self):
        """Schreibe alle Worksheets"""
        print(f"\n📊 Erstelle Excel Report...")
        
        # Summary Sheet
        self._create_summary_sheet()
        
        # Host Summary
        self._create_host_summary()
        
        # Vulnerability Sheets (nach Severity)
        for severity in ['Critical', 'High', 'Medium', 'Low', 'Info']:
            if self.parser.vulnerabilities[severity]:
                self._create_vulnerability_sheet(severity)
        
        # All Vulnerabilities
        self._create_all_vulnerabilities_sheet()
        
        # Speichern
        print(f"💾 Speichere {self.output_file.name}...")
        self.wb.save(self.output_file)
        print(f"✅ Report erstellt: {self.output_file}")
    
    def _create_summary_sheet(self):
        """Erstelle Summary Sheet"""
        ws = self.wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = 'Nessus Scan Summary'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # Scan Info
        ws[f'A{row}'] = 'Scan File:'
        ws[f'B{row}'] = self.parser.nessus_file.name
        row += 1
        
        ws[f'A{row}'] = 'Report Generated:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        
        ws[f'A{row}'] = 'Total Hosts:'
        ws[f'B{row}'] = len(self.parser.hosts)
        row += 2
        
        # Vulnerability Summary
        ws[f'A{row}'] = 'Severity'
        ws[f'B{row}'] = 'Count'
        self._apply_header_style(ws, row, 'A', 'B')
        row += 1
        
        for severity in ['Critical', 'High', 'Medium', 'Low', 'Info']:
            count = len(self.parser.vulnerabilities[severity])
            ws[f'A{row}'] = severity
            ws[f'B{row}'] = count
            
            # Severity Color
            color = NessusParser.SEVERITY_COLORS[severity]
            ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
            
            row += 1
        
        # Column width
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_host_summary(self):
        """Erstelle Host Summary Sheet"""
        ws = self.wb.create_sheet("Hosts")
        
        # Headers
        headers = ['Host Name', 'IP Address', 'FQDN', 'Operating System', 'MAC Address', 'NetBIOS Name']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            self._apply_header_style_cell(cell)
        
        # Data
        for row_idx, host in enumerate(self.parser.hosts, 2):
            ws.cell(row_idx, 1, host['name'])
            ws.cell(row_idx, 2, host['ip'])
            ws.cell(row_idx, 3, host['fqdn'])
            ws.cell(row_idx, 4, host['os'])
            ws.cell(row_idx, 5, host['mac'])
            ws.cell(row_idx, 6, host['netbios'])
        
        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        # Freeze panes
        ws.freeze_panes = 'A2'
    
    def _create_vulnerability_sheet(self, severity: str):
        """Erstelle Sheet für spezifische Severity"""
        ws = self.wb.create_sheet(severity)
        
        # Headers
        headers = [
            'Host IP', 'Host Name', 'FQDN', 'Plugin ID', 'Plugin Name', 
            'Plugin Family', 'Severity', 'Port', 'Protocol', 'Service',
            'CVSS', 'CVE', 'Synopsis', 'Description', 'Solution'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            self._apply_header_style_cell(cell)
        
        # Data
        vulns = self.parser.vulnerabilities[severity]
        for row_idx, vuln in enumerate(vulns, 2):
            ws.cell(row_idx, 1, vuln['host_ip'])
            ws.cell(row_idx, 2, vuln['host_name'])
            ws.cell(row_idx, 3, vuln['host_fqdn'])
            ws.cell(row_idx, 4, vuln['plugin_id'])
            ws.cell(row_idx, 5, vuln['plugin_name'])
            ws.cell(row_idx, 6, vuln['plugin_family'])
            ws.cell(row_idx, 7, severity)
            ws.cell(row_idx, 8, vuln['port'])
            ws.cell(row_idx, 9, vuln['protocol'])
            ws.cell(row_idx, 10, vuln['service'])
            ws.cell(row_idx, 11, vuln['cvss'])
            ws.cell(row_idx, 12, vuln['cve'])
            ws.cell(row_idx, 13, vuln['synopsis'])
            ws.cell(row_idx, 14, vuln['description'])
            ws.cell(row_idx, 15, vuln['solution'])
            
            # Severity Color
            color = NessusParser.SEVERITY_COLORS[severity]
            ws.cell(row_idx, 7).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row_idx, 7).font = Font(bold=True, color="FFFFFF")
        
        # Column widths
        widths = [15, 20, 25, 10, 35, 25, 10, 8, 10, 15, 8, 15, 40, 50, 50]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze panes
        ws.freeze_panes = 'A2'
        
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    
    def _create_all_vulnerabilities_sheet(self):
        """Erstelle Sheet mit allen Vulnerabilities"""
        ws = self.wb.create_sheet("All Vulnerabilities")
        
        # Headers
        headers = [
            'Severity', 'Host IP', 'Host Name', 'Plugin ID', 'Plugin Name', 
            'Plugin Family', 'Port', 'CVSS', 'CVE', 'Synopsis', 'Solution'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            self._apply_header_style_cell(cell)
        
        # Data - alle Vulnerabilities kombiniert
        row_idx = 2
        for severity in ['Critical', 'High', 'Medium', 'Low', 'Info']:
            for vuln in self.parser.vulnerabilities[severity]:
                ws.cell(row_idx, 1, severity)
                ws.cell(row_idx, 2, vuln['host_ip'])
                ws.cell(row_idx, 3, vuln['host_name'])
                ws.cell(row_idx, 4, vuln['plugin_id'])
                ws.cell(row_idx, 5, vuln['plugin_name'])
                ws.cell(row_idx, 6, vuln['plugin_family'])
                ws.cell(row_idx, 7, vuln['port'])
                ws.cell(row_idx, 8, vuln['cvss'])
                ws.cell(row_idx, 9, vuln['cve'])
                ws.cell(row_idx, 10, vuln['synopsis'])
                ws.cell(row_idx, 11, vuln['solution'])
                
                # Severity Color
                color = NessusParser.SEVERITY_COLORS[severity]
                ws.cell(row_idx, 1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                ws.cell(row_idx, 1).font = Font(bold=True, color="FFFFFF")
                
                row_idx += 1
        
        # Column widths
        widths = [10, 15, 20, 10, 35, 25, 8, 8, 15, 40, 50]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze panes
        ws.freeze_panes = 'A2'
        
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    
    def _apply_header_style(self, ws, row, start_col, end_col):
        """Wende Header Style auf Range an"""
        for col in range(ord(start_col), ord(end_col) + 1):
            cell = ws[f'{chr(col)}{row}']
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
    
    def _apply_header_style_cell(self, cell):
        """Wende Header Style auf einzelne Cell an"""
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.border


def process_file(nessus_file: Path, output_file: Path = None) -> bool:
    """Verarbeite eine einzelne Nessus Datei"""
    
    # Output Dateiname
    if output_file is None:
        output_file = nessus_file.with_suffix('.xlsx')
    
    # Parse Nessus
    parser = NessusParser(nessus_file)
    if not parser.parse():
        return False
    
    # Schreibe Excel
    writer = ExcelWriter(output_file, parser)
    writer.write()
    
    return True


def process_directory(directory: Path, pattern: str = "*.nessus"):
    """Verarbeite alle Nessus Dateien in einem Verzeichnis"""
    
    nessus_files = list(directory.glob(pattern))
    
    if not nessus_files:
        print(f"❌ Keine .nessus Dateien in {directory} gefunden!")
        return
    
    print(f"📂 {len(nessus_files)} Nessus Dateien gefunden\n")
    
    success_count = 0
    fail_count = 0
    
    for nessus_file in nessus_files:
        print(f"\n{'='*60}")
        if process_file(nessus_file):
            success_count += 1
        else:
            fail_count += 1
    
    print(f"\n{'='*60}")
    print(f"✅ Erfolgreich: {success_count}")
    if fail_count > 0:
        print(f"❌ Fehler: {fail_count}")


def main():
    parser = argparse.ArgumentParser(
        description='Konvertiert Nessus XML Dateien zu Excel XLSX',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Beispiele:
  %(prog)s scan.nessus
  %(prog)s scan.nessus -o report.xlsx
  %(prog)s --dir /pfad/zu/scans
  %(prog)s --dir /pfad/zu/scans --pattern "*.xml"
        """
    )
    
    parser.add_argument(
        'input',
        nargs='?',
        help='Nessus Datei (.nessus oder .xml)'
    )
    
    parser.add_argument(
        '--dir', '-d',
        type=Path,
        help='Verzeichnis mit Nessus Dateien'
    )
    
    parser.add_argument(
        '--output', '-o',
        type=Path,
        help='Output Excel Datei (nur bei einzelner Datei)'
    )
    
    parser.add_argument(
        '--pattern', '-p',
        default='*.nessus',
        help='Datei Pattern für Verzeichnis-Modus (default: *.nessus)'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Zeige detaillierte Debug-Informationen'
    )
    
    args = parser.parse_args()
    
    # Prüfe Abhängigkeiten
    try:
        import openpyxl
    except ImportError:
        print("❌ openpyxl ist nicht installiert!")
        print("   Installiere mit: pip install openpyxl")
        sys.exit(1)
    
    # Directory oder File Mode
    if args.dir:
        if not args.dir.is_dir():
            print(f"❌ Verzeichnis nicht gefunden: {args.dir}")
            sys.exit(1)
        process_directory(args.dir, args.pattern)
    
    elif args.input:
        input_file = Path(args.input)
        if not input_file.is_file():
            print(f"❌ Datei nicht gefunden: {input_file}")
            sys.exit(1)
        
        if not process_file(input_file, args.output):
            sys.exit(1)
    
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == '__main__':
    main()
