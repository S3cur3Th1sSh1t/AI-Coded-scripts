#!/usr/bin/env python3
"""
PyADRecon - Python Active Directory Reconnaissance Tool
A Python port of ADRecon with NTLM and Kerberos authentication support.

Author: S3cur3Th1sSh1t, LRVT
License: MIT
"""

import argparse
import csv
import os
import sys
import socket
import struct
import ssl
import re
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from pathlib import Path
import logging
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

# Third-party imports
LDAP3_AVAILABLE = False
# Default values if ldap3 not available (for --help to work)
SUBTREE = 2
ALL_ATTRIBUTES = '*'
LDAPException = Exception
LDAPBindError = Exception

try:
    import ldap3
    from ldap3 import Server, Connection, ALL, NTLM, KERBEROS, SASL, SUBTREE, ALL_ATTRIBUTES
    from ldap3.core.exceptions import LDAPException, LDAPBindError
    from ldap3.utils.conv import escape_filter_chars
    LDAP3_AVAILABLE = True
except ImportError:
    pass

try:
    from impacket.krb5.kerberosv5 import getKerberosTGT, getKerberosTGS
    from impacket.krb5.types import Principal, KerberosTime, Ticket
    from impacket.krb5 import constants
    from impacket.krb5.asn1 import TGS_REP, EncTGSRepPart, EncTicketPart
    from impacket.krb5.ccache import CCache
    from impacket.smbconnection import SMBConnection
    from impacket.nmb import NetBIOSError
    IMPACKET_AVAILABLE = True
except ImportError:
    IMPACKET_AVAILABLE = False
    print("[*] impacket not available - Kerberoast and SMB features disabled")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[*] openpyxl not available - Excel export disabled")


# Constants
VERSION = "1.0.0"
BANNER = f"""
╔═══════════════════════════════════════════════════════════╗
║  PyADRecon v{VERSION} - Python AD Reconnaissance Tool      ║
║  A Python implementation inspired by ADRecon              ║
╚═══════════════════════════════════════════════════════════╝
"""

# AD Constants
SAM_USER_OBJECT = 805306368
SAM_COMPUTER_OBJECT = 805306369
SAM_GROUP_OBJECT = 268435456
SAM_ALIAS_OBJECT = 536870912

# UserAccountControl flags
UAC_FLAGS = {
    0x0001: "SCRIPT",
    0x0002: "ACCOUNTDISABLE",
    0x0008: "HOMEDIR_REQUIRED",
    0x0010: "LOCKOUT",
    0x0020: "PASSWD_NOTREQD",
    0x0040: "PASSWD_CANT_CHANGE",
    0x0080: "ENCRYPTED_TEXT_PWD_ALLOWED",
    0x0100: "TEMP_DUPLICATE_ACCOUNT",
    0x0200: "NORMAL_ACCOUNT",
    0x0800: "INTERDOMAIN_TRUST_ACCOUNT",
    0x1000: "WORKSTATION_TRUST_ACCOUNT",
    0x2000: "SERVER_TRUST_ACCOUNT",
    0x10000: "DONT_EXPIRE_PASSWORD",
    0x20000: "MNS_LOGON_ACCOUNT",
    0x40000: "SMARTCARD_REQUIRED",
    0x80000: "TRUSTED_FOR_DELEGATION",
    0x100000: "NOT_DELEGATED",
    0x200000: "USE_DES_KEY_ONLY",
    0x400000: "DONT_REQ_PREAUTH",
    0x800000: "PASSWORD_EXPIRED",
    0x1000000: "TRUSTED_TO_AUTH_FOR_DELEGATION",
    0x4000000: "PARTIAL_SECRETS_ACCOUNT",
}

# Kerberos encryption types
KERB_ENC_FLAGS = {
    0x01: "DES_CBC_CRC",
    0x02: "DES_CBC_MD5",
    0x04: "RC4_HMAC",
    0x08: "AES128_CTS_HMAC_SHA1_96",
    0x10: "AES256_CTS_HMAC_SHA1_96",
}

# Trust directions
TRUST_DIRECTION = {
    0: "Disabled",
    1: "Inbound",
    2: "Outbound",
    3: "Bidirectional",
}

# Trust types
TRUST_TYPE = {
    1: "Downlevel",
    2: "Uplevel",
    3: "MIT",
    4: "DCE",
}

# Domain functional levels
DOMAIN_FUNCTIONAL_LEVELS = {
    0: "Windows2000",
    1: "Windows2003Mixed",
    2: "Windows2003",
    3: "Windows2008",
    4: "Windows2008R2",
    5: "Windows2012",
    6: "Windows2012R2",
    7: "Windows2016",
}

# Group types
GROUP_TYPE = {
    2: "Global Distribution",
    4: "Domain Local Distribution",
    8: "Universal Distribution",
    -2147483646: "Global Security",
    -2147483644: "Domain Local Security",
    -2147483640: "Universal Security",
}

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format='[%(levelname)s] %(message)s'
)
logger = logging.getLogger('PyADRecon')


def windows_timestamp_to_datetime(timestamp: int) -> Optional[datetime]:
    """Convert Windows FILETIME to Python datetime."""
    if timestamp is None or timestamp == 0 or timestamp == 9223372036854775807:
        return None
    try:
        # Windows FILETIME is 100-nanosecond intervals since January 1, 1601
        return datetime(1601, 1, 1) + timedelta(microseconds=timestamp // 10)
    except (ValueError, OverflowError):
        return None


def generalized_time_to_datetime(time_str: str) -> Optional[datetime]:
    """Convert LDAP Generalized Time to Python datetime."""
    if not time_str:
        return None
    try:
        # Handle various formats
        if time_str.endswith('Z'):
            time_str = time_str[:-1]
        if '.' in time_str:
            return datetime.strptime(time_str, "%Y%m%d%H%M%S.%f")
        return datetime.strptime(time_str, "%Y%m%d%H%M%S")
    except ValueError:
        return None


def format_datetime(dt) -> str:
    """Format datetime to match ADRecon output format (M/D/YYYY H:MM:SS AM/PM)."""
    if dt is None:
        return ""
    # If it's already a datetime object from ldap3
    if isinstance(dt, datetime):
        return dt.strftime("%-m/%-d/%Y %-I:%M:%S %p")
    return ""


def _extract_ldap_value(attr):
    """Extract primitive value from ldap3 Attribute object."""
    # ldap3 Attribute objects have 'raw_values' attribute - use this to detect them
    if hasattr(attr, 'raw_values'):
        # It's an ldap3 Attribute object
        return attr.value
    return attr


def safe_int(val, default=0):
    """Safely convert a value to int, handling ldap3 Attribute objects."""
    if val is None:
        return default
    # Handle ldap3 Attribute objects
    if hasattr(val, 'raw_values'):
        val = val.value
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def safe_str(val, default=''):
    """Safely convert a value to string, handling ldap3 Attribute objects."""
    if val is None:
        return default
    # Handle ldap3 Attribute objects
    if hasattr(val, 'raw_values'):
        val = val.value
    if val is None:
        return default
    return str(val)


def get_attr(entry, attr_name: str, default=None):
    """Safely get attribute value from LDAP entry."""
    try:
        if hasattr(entry, attr_name):
            attr = getattr(entry, attr_name)
            if attr is not None:
                # Extract value from ldap3 Attribute if needed
                val = _extract_ldap_value(attr)
                if val is not None:
                    if isinstance(val, list):
                        return val[0] if len(val) == 1 else val
                    return val
    except (IndexError, KeyError, AttributeError):
        pass
    return default


def get_attr_list(entry, attr_name: str) -> List:
    """Get attribute as a list."""
    try:
        if hasattr(entry, attr_name):
            attr = getattr(entry, attr_name)
            if attr is not None:
                # ldap3 Attribute objects have 'raw_values' attribute
                if hasattr(attr, 'raw_values'):
                    vals = attr.values
                    if vals:
                        return list(vals)
                    return []
                elif isinstance(attr, list):
                    return attr
                else:
                    return [attr]
    except (IndexError, KeyError, AttributeError):
        pass
    return []


def dn_to_fqdn(dn: str) -> str:
    """Convert Distinguished Name to FQDN."""
    if not dn:
        return ""
    parts = []
    for part in dn.split(','):
        if part.upper().startswith('DC='):
            parts.append(part[3:])
    return '.'.join(parts)


def parse_uac(uac) -> Dict[str, bool]:
    """Parse UserAccountControl value into individual flags."""
    if uac is None:
        return {}
    uac = safe_int(uac, 0)
    result = {}
    result['Enabled'] = not bool(uac & 0x0002)
    result['PasswordNotRequired'] = bool(uac & 0x0020)
    result['PasswordCantChange'] = bool(uac & 0x0040)
    result['PasswordNeverExpires'] = bool(uac & 0x10000)
    result['SmartcardRequired'] = bool(uac & 0x40000)
    result['TrustedForDelegation'] = bool(uac & 0x80000)
    result['NotDelegated'] = bool(uac & 0x100000)
    result['UseDESKeyOnly'] = bool(uac & 0x200000)
    result['DoesNotRequirePreAuth'] = bool(uac & 0x400000)
    result['PasswordExpired'] = bool(uac & 0x800000)
    result['TrustedToAuthForDelegation'] = bool(uac & 0x1000000)
    result['AccountLockedOut'] = bool(uac & 0x0010)
    return result


def parse_kerb_enc_types(enc_types) -> Dict[str, bool]:
    """Parse msDS-SupportedEncryptionTypes into individual encryption types."""
    if enc_types is None:
        return {}
    enc_types = safe_int(enc_types, 0)
    return {
        'RC4': bool(enc_types & 0x04),
        'AES128': bool(enc_types & 0x08),
        'AES256': bool(enc_types & 0x10),
    }


def sid_to_string(sid_bytes) -> str:
    """Convert binary SID to string representation."""
    if sid_bytes is None:
        return ""
    if isinstance(sid_bytes, str):
        return sid_bytes
    try:
        if isinstance(sid_bytes, bytes):
            revision = sid_bytes[0]
            sub_auth_count = sid_bytes[1]
            authority = int.from_bytes(sid_bytes[2:8], byteorder='big')
            sub_auths = []
            for i in range(sub_auth_count):
                sub_auth = struct.unpack('<I', sid_bytes[8 + 4*i:12 + 4*i])[0]
                sub_auths.append(str(sub_auth))
            return f"S-{revision}-{authority}-" + '-'.join(sub_auths)
    except Exception:
        pass
    return str(sid_bytes)


@dataclass
class ADReconConfig:
    """Configuration for AD reconnaissance."""
    domain_controller: str
    domain: str = ""
    username: str = ""
    password: str = ""
    auth_method: str = "ntlm"  # ntlm, kerberos
    use_ssl: bool = False
    port: int = 389
    page_size: int = 500
    threads: int = 10
    dormant_days: int = 90
    password_age_days: int = 30
    output_dir: str = ""
    only_enabled: bool = False

    # Collection flags
    collect_forest: bool = True
    collect_domain: bool = True
    collect_trusts: bool = True
    collect_sites: bool = True
    collect_subnets: bool = True
    collect_schema: bool = True
    collect_password_policy: bool = True
    collect_fgpp: bool = True
    collect_dcs: bool = True
    collect_users: bool = True
    collect_user_spns: bool = True
    collect_groups: bool = True
    collect_group_members: bool = True
    collect_ous: bool = True
    collect_gpos: bool = True
    collect_gplinks: bool = True
    collect_dns_zones: bool = True
    collect_dns_records: bool = True
    collect_printers: bool = True
    collect_computers: bool = True
    collect_computer_spns: bool = True
    collect_laps: bool = True
    collect_bitlocker: bool = True
    collect_kerberoast: bool = False
    collect_acls: bool = False


class PyADRecon:
    """Main AD Reconnaissance class."""

    def __init__(self, config: ADReconConfig):
        self.config = config
        self.conn: Optional[Connection] = None
        self.base_dn: str = ""
        self.config_dn: str = ""
        self.schema_dn: str = ""
        self.domain_sid: str = ""
        self.results: Dict[str, List] = {}
        self.start_time: datetime = datetime.now()

    def connect(self) -> bool:
        """Establish LDAP connection."""
        try:
            port = 636 if self.config.use_ssl else self.config.port
            server = Server(
                self.config.domain_controller,
                port=port,
                use_ssl=self.config.use_ssl,
                get_info=ALL
            )

            if self.config.auth_method.lower() == 'kerberos':
                logger.info("Connecting using Kerberos authentication...")
                self.conn = Connection(
                    server,
                    user=self.config.username,
                    password=self.config.password,
                    authentication=SASL,
                    sasl_mechanism=KERBEROS,
                    auto_bind=True
                )
            else:
                # NTLM authentication
                logger.info("Connecting using NTLM authentication...")
                user = self.config.username
                if '\\' not in user and '@' not in user:
                    if self.config.domain:
                        user = f"{self.config.domain}\\{user}"

                self.conn = Connection(
                    server,
                    user=user,
                    password=self.config.password,
                    authentication=NTLM,
                    auto_bind=True
                )

            if self.conn.bound:
                logger.info("LDAP bind successful")
                self._get_root_dse()
                return True
            else:
                logger.error(f"LDAP bind failed: {self.conn.result}")
                return False

        except LDAPBindError as e:
            logger.error(f"LDAP bind error: {e}")
            return False
        except LDAPException as e:
            logger.error(f"LDAP error: {e}")
            return False
        except Exception as e:
            logger.error(f"Connection error: {e}")
            return False

    def _get_root_dse(self):
        """Get root DSE information."""
        if self.conn.server.info:
            info = self.conn.server.info
            if info.naming_contexts:
                self.base_dn = str(info.naming_contexts[0])
            if hasattr(info, 'other'):
                if 'configurationNamingContext' in info.other:
                    self.config_dn = str(info.other['configurationNamingContext'][0])
                if 'schemaNamingContext' in info.other:
                    self.schema_dn = str(info.other['schemaNamingContext'][0])
                if 'defaultNamingContext' in info.other:
                    self.base_dn = str(info.other['defaultNamingContext'][0])

        logger.info(f"Base DN: {self.base_dn}")
        logger.info(f"Config DN: {self.config_dn}")

    def search(self, search_base: str, search_filter: str, attributes: List[str] = None,
               search_scope: int = SUBTREE) -> List:
        """Perform paged LDAP search."""
        if attributes is None:
            attributes = ['*']

        entries = []
        try:
            self.conn.search(
                search_base=search_base,
                search_filter=search_filter,
                search_scope=search_scope,
                attributes=attributes,
                paged_size=self.config.page_size,
                paged_cookie=None
            )

            entries.extend(self.conn.entries)

            # Handle paging
            while self.conn.result.get('controls', {}).get('1.2.840.113556.1.4.319', {}).get('value', {}).get('cookie'):
                cookie = self.conn.result['controls']['1.2.840.113556.1.4.319']['value']['cookie']
                self.conn.search(
                    search_base=search_base,
                    search_filter=search_filter,
                    search_scope=search_scope,
                    attributes=attributes,
                    paged_size=self.config.page_size,
                    paged_cookie=cookie
                )
                entries.extend(self.conn.entries)

        except LDAPException as e:
            logger.warning(f"Search error: {e}")

        return entries

    def collect_domain_info(self) -> List[Dict]:
        """Collect domain information."""
        logger.info("[-] Collecting Domain Information...")
        results = []

        try:
            # Query domain object
            # Using objectCategory=domainDNS instead of objectClass=domain
            entries = self.search(
                self.base_dn,
                "(objectCategory=domainDNS)",
                ['*']
            )

            if entries:
                entry = entries[0]

                # Get domain functional level
                fl = get_attr(entry, 'msDS-Behavior-Version', 0)
                func_level = DOMAIN_FUNCTIONAL_LEVELS.get(safe_int(fl), "Unknown")

                # Get domain SID
                sid_bytes = get_attr(entry, 'objectSid')
                self.domain_sid = sid_to_string(sid_bytes)

                results.append({"Category": "Name", "Value": dn_to_fqdn(self.base_dn)})
                results.append({"Category": "NetBIOS", "Value": get_attr(entry, 'name', '')})
                results.append({"Category": "Functional Level", "Value": f"{func_level}Domain"})
                results.append({"Category": "DomainSID", "Value": self.domain_sid})
                results.append({"Category": "Creation Date", "Value": format_datetime(get_attr(entry, 'whenCreated'))})
                results.append({"Category": "ms-DS-MachineAccountQuota", "Value": str(get_attr(entry, 'ms-DS-MachineAccountQuota', ''))})

                # Get RID info
                rid_entries = self.search(
                    f"CN=RID Manager$,CN=System,{self.base_dn}",
                    "(objectClass=*)",
                    ['rIDAvailablePool']
                )
                if rid_entries:
                    rid_pool = get_attr(rid_entries[0], 'rIDAvailablePool')
                    if rid_pool:
                        rid_pool = safe_int(rid_pool)
                        total_sids = rid_pool >> 32
                        issued = rid_pool & 0xFFFFFFFF
                        remaining = total_sids - issued
                        results.append({"Category": "RIDs Issued", "Value": str(issued)})
                        results.append({"Category": "RIDs Remaining", "Value": str(remaining)})

        except Exception as e:
            logger.warning(f"Error collecting domain info: {e}")

        self.results['Domain'] = results
        logger.info(f"    Found {len(results)} domain properties")
        return results

    def collect_forest_info(self) -> List[Dict]:
        """Collect forest information."""
        logger.info("[-] Collecting Forest Information...")
        results = []

        try:
            # Get forest info from configuration partition
            entries = self.search(
                f"CN=Partitions,{self.config_dn}",
                "(objectClass=crossRefContainer)",
                ['*']
            )

            if entries:
                entry = entries[0]
                forest_fl = get_attr(entry, 'msDS-Behavior-Version', 0)
                func_level = DOMAIN_FUNCTIONAL_LEVELS.get(safe_int(forest_fl), "Unknown")

                results.append({"Category": "Name", "Value": dn_to_fqdn(self.base_dn)})
                results.append({"Category": "Functional Level", "Value": f"{func_level}Forest"})

            # Check for LAPS
            laps_entries = self.search(
                self.schema_dn,
                "(name=ms-Mcs-AdmPwd)",
                ['whenCreated']
            )
            if laps_entries:
                results.append({"Category": "LAPS", "Value": "Enabled"})
                results.append({"Category": "LAPS Installed Date", "Value": format_datetime(get_attr(laps_entries[0], 'whenCreated'))})
            else:
                results.append({"Category": "LAPS", "Value": "Not Installed"})

        except Exception as e:
            logger.warning(f"Error collecting forest info: {e}")

        self.results['Forest'] = results
        logger.info(f"    Found {len(results)} forest properties")
        return results

    def collect_trusts(self) -> List[Dict]:
        """Collect trust relationships."""
        logger.info("[-] Collecting Trust Relationships...")
        results = []

        try:
            # Using objectCategory=trustedDomain for slightly different query
            entries = self.search(
                self.base_dn,
                "(objectCategory=trustedDomain)",
                ['distinguishedName', 'trustPartner', 'trustDirection', 'trustType',
                 'trustAttributes', 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                trust_dir = get_attr(entry, 'trustDirection', 0)
                trust_type = get_attr(entry, 'trustType', 0)
                trust_attrs = get_attr(entry, 'trustAttributes', 0)

                # Parse trust attributes
                attrs_list = []
                if trust_attrs:
                    trust_attrs = safe_int(trust_attrs)
                    if trust_attrs & 0x01: attrs_list.append("Non Transitive")
                    if trust_attrs & 0x02: attrs_list.append("UpLevel")
                    if trust_attrs & 0x04: attrs_list.append("Quarantined")
                    if trust_attrs & 0x08: attrs_list.append("Forest Transitive")
                    if trust_attrs & 0x10: attrs_list.append("Cross Organization")
                    if trust_attrs & 0x20: attrs_list.append("Within Forest")
                    if trust_attrs & 0x40: attrs_list.append("Treat as External")
                    if trust_attrs & 0x80: attrs_list.append("Uses RC4 Encryption")
                    if trust_attrs & 0x200: attrs_list.append("No TGT Delegation")

                results.append({
                    "Source Domain": dn_to_fqdn(str(get_attr(entry, 'distinguishedName', ''))),
                    "Target Domain": get_attr(entry, 'trustPartner', ''),
                    "Trust Direction": TRUST_DIRECTION.get(safe_int(trust_dir), "Unknown"),
                    "Trust Type": TRUST_TYPE.get(safe_int(trust_type), "Unknown"),
                    "Attributes": ", ".join(attrs_list),
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                    "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                })

        except Exception as e:
            logger.warning(f"Error collecting trusts: {e}")

        self.results['Trusts'] = results
        logger.info(f"    Found {len(results)} trust relationships")
        return results

    def collect_sites(self) -> List[Dict]:
        """Collect AD sites."""
        logger.info("[-] Collecting Sites...")
        results = []

        try:
            entries = self.search(
                f"CN=Sites,{self.config_dn}",
                "(objectCategory=site)",
                ['name', 'description', 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                results.append({
                    "Name": get_attr(entry, 'name', ''),
                    "Description": get_attr(entry, 'description', ''),
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                    "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                })

        except Exception as e:
            logger.warning(f"Error collecting sites: {e}")

        self.results['Sites'] = results
        logger.info(f"    Found {len(results)} sites")
        return results

    def collect_subnets(self) -> List[Dict]:
        """Collect AD subnets."""
        logger.info("[-] Collecting Subnets...")
        results = []

        try:
            entries = self.search(
                f"CN=Subnets,CN=Sites,{self.config_dn}",
                "(objectCategory=subnet)",
                ['name', 'description', 'siteObject', 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                site_obj = get_attr(entry, 'siteObject', '')
                site_name = ""
                if site_obj:
                    # Extract site name from DN
                    match = re.search(r'CN=([^,]+)', str(site_obj))
                    if match:
                        site_name = match.group(1)

                results.append({
                    "Site": site_name,
                    "Name": get_attr(entry, 'name', ''),
                    "Description": get_attr(entry, 'description', ''),
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                    "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                })

        except Exception as e:
            logger.warning(f"Error collecting subnets: {e}")

        self.results['Subnets'] = results
        logger.info(f"    Found {len(results)} subnets")
        return results

    def collect_domain_controllers(self) -> List[Dict]:
        """Collect domain controller information."""
        logger.info("[-] Collecting Domain Controllers...")
        results = []

        try:
            # Find DCs using primaryGroupID or userAccountControl
            # Using (&(objectCategory=computer)(userAccountControl:1.2.840.113556.1.4.803:=8192))
            entries = self.search(
                self.base_dn,
                "(&(objectCategory=computer)(userAccountControl:1.2.840.113556.1.4.803:=8192))",
                ['name', 'dNSHostName', 'operatingSystem', 'operatingSystemVersion',
                 'operatingSystemServicePack', 'whenCreated', 'whenChanged',
                 'msDS-isRODC', 'serverReferenceBL']
            )

            for entry in entries:
                is_rodc = get_attr(entry, 'msDS-isRODC', False)

                dc_info = {
                    "Name": safe_str(get_attr(entry, 'name', '')),
                    "Hostname": safe_str(get_attr(entry, 'dNSHostName', '')),
                    "Operating System": safe_str(get_attr(entry, 'operatingSystem', '')),
                    "OS Version": safe_str(get_attr(entry, 'operatingSystemVersion', '')),
                    "OS Service Pack": safe_str(get_attr(entry, 'operatingSystemServicePack', '')),
                    "Is RODC": str(is_rodc),
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                    "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                }

                # Try to get IP address
                hostname = safe_str(get_attr(entry, 'dNSHostName', ''))
                if hostname:
                    try:
                        ip = socket.gethostbyname(hostname)
                        dc_info["IPv4Address"] = ip
                    except socket.gaierror:
                        dc_info["IPv4Address"] = ""

                results.append(dc_info)

        except Exception as e:
            logger.warning(f"Error collecting DCs: {e}")

        self.results['DomainControllers'] = results
        logger.info(f"    Found {len(results)} domain controllers")
        return results

    def collect_password_policy(self) -> List[Dict]:
        """Collect default domain password policy."""
        logger.info("[-] Collecting Default Password Policy...")
        results = []

        try:
            entries = self.search(
                self.base_dn,
                "(objectCategory=domainDNS)",
                ['minPwdLength', 'pwdHistoryLength', 'pwdProperties',
                 'maxPwdAge', 'minPwdAge', 'lockoutThreshold',
                 'lockoutDuration', 'lockoutObservationWindow']
            )

            if entries:
                entry = entries[0]

                # Convert password age from 100-nanosecond intervals to days
                max_pwd_age = get_attr(entry, 'maxPwdAge')
                min_pwd_age = get_attr(entry, 'minPwdAge')
                lockout_duration = get_attr(entry, 'lockoutDuration')
                lockout_window = get_attr(entry, 'lockoutObservationWindow')

                def convert_interval_to_days(interval):
                    if interval is None:
                        return "Not Set"
                    try:
                        # Negative value in 100-nanosecond intervals
                        interval = abs(safe_int(interval))
                        if interval == 0:
                            return "Not Set"
                        days = interval / (10000000 * 60 * 60 * 24)
                        return f"{days:.2f} days"
                    except:
                        return str(interval)

                def convert_interval_to_minutes(interval):
                    if interval is None:
                        return "Not Set"
                    try:
                        interval = abs(safe_int(interval))
                        if interval == 0:
                            return "Not Set"
                        minutes = interval / (10000000 * 60)
                        return f"{minutes:.0f} minutes"
                    except:
                        return str(interval)

                pwd_props = get_attr(entry, 'pwdProperties', 0)
                pwd_props = safe_int(pwd_props)

                results.append({"Policy": "Minimum Password Length", "Value": str(get_attr(entry, 'minPwdLength', ''))})
                results.append({"Policy": "Password History Length", "Value": str(get_attr(entry, 'pwdHistoryLength', ''))})
                results.append({"Policy": "Maximum Password Age", "Value": convert_interval_to_days(max_pwd_age)})
                results.append({"Policy": "Minimum Password Age", "Value": convert_interval_to_days(min_pwd_age)})
                results.append({"Policy": "Lockout Threshold", "Value": str(get_attr(entry, 'lockoutThreshold', ''))})
                results.append({"Policy": "Lockout Duration", "Value": convert_interval_to_minutes(lockout_duration)})
                results.append({"Policy": "Lockout Observation Window", "Value": convert_interval_to_minutes(lockout_window)})
                results.append({"Policy": "Password Complexity Required", "Value": str(bool(pwd_props & 1))})
                results.append({"Policy": "Reversible Encryption Enabled", "Value": str(bool(pwd_props & 16))})

        except Exception as e:
            logger.warning(f"Error collecting password policy: {e}")

        self.results['PasswordPolicy'] = results
        logger.info(f"    Found {len(results)} password policy settings")
        return results

    def collect_fine_grained_password_policies(self) -> List[Dict]:
        """Collect fine-grained password policies."""
        logger.info("[-] Collecting Fine-Grained Password Policies...")
        results = []

        try:
            entries = self.search(
                f"CN=Password Settings Container,CN=System,{self.base_dn}",
                "(objectCategory=msDS-PasswordSettings)",
                ['name', 'msDS-PasswordSettingsPrecedence', 'msDS-MinimumPasswordLength',
                 'msDS-PasswordHistoryLength', 'msDS-PasswordComplexityEnabled',
                 'msDS-PasswordReversibleEncryptionEnabled', 'msDS-MaximumPasswordAge',
                 'msDS-MinimumPasswordAge', 'msDS-LockoutThreshold', 'msDS-LockoutDuration',
                 'msDS-LockoutObservationWindow', 'msDS-PSOAppliesTo']
            )

            for entry in entries:
                applies_to = get_attr_list(entry, 'msDS-PSOAppliesTo')

                results.append({
                    "Name": get_attr(entry, 'name', ''),
                    "Precedence": str(get_attr(entry, 'msDS-PasswordSettingsPrecedence', '')),
                    "Min Password Length": str(get_attr(entry, 'msDS-MinimumPasswordLength', '')),
                    "Password History": str(get_attr(entry, 'msDS-PasswordHistoryLength', '')),
                    "Complexity Enabled": str(get_attr(entry, 'msDS-PasswordComplexityEnabled', '')),
                    "Reversible Encryption": str(get_attr(entry, 'msDS-PasswordReversibleEncryptionEnabled', '')),
                    "Lockout Threshold": str(get_attr(entry, 'msDS-LockoutThreshold', '')),
                    "Applies To": ", ".join([str(a) for a in applies_to]) if applies_to else "",
                })

        except Exception as e:
            logger.warning(f"Error collecting FGPP: {e}")

        self.results['FineGrainedPasswordPolicy'] = results
        logger.info(f"    Found {len(results)} fine-grained password policies")
        return results

    def collect_users(self) -> List[Dict]:
        """Collect user objects."""
        logger.info("[-] Collecting Users - May take some time...")
        results = []
        now = datetime.now()

        try:
            # Using (&(objectCategory=person)(objectClass=user)) instead of samAccountType
            filter_str = "(&(objectCategory=person)(objectClass=user))"
            if self.config.only_enabled:
                filter_str = "(&(objectCategory=person)(objectClass=user)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"

            entries = self.search(
                self.base_dn,
                filter_str,
                ['sAMAccountName', 'name', 'distinguishedName', 'canonicalName',
                 'userAccountControl', 'pwdLastSet', 'lastLogonTimestamp',
                 'accountExpires', 'adminCount', 'description', 'title',
                 'department', 'company', 'manager', 'mail', 'mobile',
                 'homeDirectory', 'profilePath', 'scriptPath', 'memberOf',
                 'primaryGroupID', 'objectSid', 'sIDHistory', 'servicePrincipalName',
                 'msDS-AllowedToDelegateTo', 'msDS-SupportedEncryptionTypes',
                 'givenName', 'sn', 'middleName', 'c', 'info', 'logonWorkstation',
                 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                uac = get_attr(entry, 'userAccountControl', 0)
                uac_parsed = parse_uac(uac)

                # Password last set
                pwd_last_set = get_attr(entry, 'pwdLastSet')
                # ldap3 returns datetime objects for these attributes, not integers
                pwd_last_set_dt = None
                pwd_age_days = None
                must_change_pwd = False
                pwd_not_changed_max = False
                
                if isinstance(pwd_last_set, datetime):
                    # Convert timezone-aware datetime to naive for comparison
                    pwd_last_set_dt = pwd_last_set.replace(tzinfo=None) if pwd_last_set.tzinfo else pwd_last_set
                    # Check if pwdLastSet is 0 (1601-01-01 epoch) which means "must change password at logon"
                    if pwd_last_set_dt.year == 1601:
                        must_change_pwd = True
                        pwd_last_set_dt = None  # Don't show the 1601 date
                    else:
                        pwd_age_days = (now - pwd_last_set_dt).days
                        if pwd_age_days > self.config.password_age_days:
                            pwd_not_changed_max = True
                else:
                    # If not a datetime, user must change password at next logon
                    must_change_pwd = True

                # Last logon
                last_logon = get_attr(entry, 'lastLogonTimestamp')
                # ldap3 returns datetime objects for these attributes, not integers
                last_logon_dt = None
                logon_age_days = None
                never_logged_in = True
                dormant = False

                if isinstance(last_logon, datetime):
                    # Convert timezone-aware datetime to naive for comparison
                    last_logon_dt = last_logon.replace(tzinfo=None) if last_logon.tzinfo else last_logon
                    # Check if lastLogonTimestamp is 0 (1601-01-01 epoch) which means never logged in
                    if last_logon_dt.year == 1601:
                        last_logon_dt = None
                        never_logged_in = True
                    else:
                        never_logged_in = False
                        logon_age_days = (now - last_logon_dt).days
                        if logon_age_days > self.config.dormant_days:
                            dormant = True

                # Account expiration
                acc_expires = get_attr(entry, 'accountExpires')
                acc_expires_int = safe_int(acc_expires)
                acc_expires_dt = windows_timestamp_to_datetime(acc_expires_int) if acc_expires_int else None
                acc_expires_days = None
                if acc_expires_dt:
                    acc_expires_days = (acc_expires_dt - now).days

                # Kerberos encryption types
                enc_types = get_attr(entry, 'msDS-SupportedEncryptionTypes')
                kerb_enc = parse_kerb_enc_types(enc_types)

                # Delegation
                delegation_type = None
                delegation_protocol = None
                delegation_services = None

                if uac_parsed.get('TrustedForDelegation', False):
                    delegation_type = "Unconstrained"
                    delegation_services = "Any"

                allowed_to_delegate = get_attr_list(entry, 'msDS-AllowedToDelegateTo')
                if allowed_to_delegate:
                    delegation_type = "Constrained"
                    delegation_services = ", ".join([str(s) for s in allowed_to_delegate])

                if uac_parsed.get('TrustedToAuthForDelegation', False):
                    delegation_protocol = "Any"
                elif delegation_type:
                    delegation_protocol = "Kerberos"

                # SPNs
                spns = get_attr_list(entry, 'servicePrincipalName')
                has_spn = len(spns) > 0

                # SID History
                sid_history = get_attr_list(entry, 'sIDHistory')
                sid_history_str = ", ".join([sid_to_string(s) for s in sid_history]) if sid_history else ""

                results.append({
                    "UserName": get_attr(entry, 'sAMAccountName', ''),
                    "Name": get_attr(entry, 'name', ''),
                    "Enabled": uac_parsed.get('Enabled', ''),
                    "Must Change Password at Logon": must_change_pwd,
                    "Cannot Change Password": uac_parsed.get('PasswordCantChange', False),
                    "Password Never Expires": uac_parsed.get('PasswordNeverExpires', False),
                    "Reversible Password Encryption": False,  # Would need to check AD policies
                    "Smartcard Logon Required": uac_parsed.get('SmartcardRequired', False),
                    "Delegation Permitted": not uac_parsed.get('NotDelegated', False),
                    "Kerberos DES Only": uac_parsed.get('UseDESKeyOnly', False),
                    "Kerberos RC4": kerb_enc.get('RC4', '') if kerb_enc.get('RC4') else '',
                    "Kerberos AES-128bit": kerb_enc.get('AES128', '') if kerb_enc.get('AES128') else '',
                    "Kerberos AES-256bit": kerb_enc.get('AES256', '') if kerb_enc.get('AES256') else '',
                    "Does Not Require Pre Auth": uac_parsed.get('DoesNotRequirePreAuth', False),
                    "Never Logged in": never_logged_in,
                    "Logon Age (days)": logon_age_days if logon_age_days is not None else "",
                    "Password Age (days)": pwd_age_days if pwd_age_days is not None else "",
                    f"Dormant (> {self.config.dormant_days} days)": dormant,
                    f"Password Age (> {self.config.password_age_days} days)": pwd_not_changed_max,
                    "Account Locked Out": uac_parsed.get('AccountLockedOut', False),
                    "Password Expired": uac_parsed.get('PasswordExpired', False),
                    "Password Not Required": uac_parsed.get('PasswordNotRequired', False),
                    "Delegation Type": delegation_type or "",
                    "Delegation Protocol": delegation_protocol or "",
                    "Delegation Services": delegation_services or "",
                    "Logon Workstations": get_attr(entry, 'logonWorkstation', ''),
                    "AdminCount": get_attr(entry, 'adminCount', ''),
                    "Primary GroupID": get_attr(entry, 'primaryGroupID', ''),
                    "SID": sid_to_string(get_attr(entry, 'objectSid')),
                    "SIDHistory": sid_history_str,
                    "HasSPN": has_spn,
                    "Description": get_attr(entry, 'description', ''),
                    "Title": get_attr(entry, 'title', ''),
                    "Department": get_attr(entry, 'department', ''),
                    "Company": get_attr(entry, 'company', ''),
                    "Manager": get_attr(entry, 'manager', ''),
                    "Info": get_attr(entry, 'info', ''),
                    "Last Logon Date": format_datetime(last_logon_dt),
                    "Password LastSet": format_datetime(pwd_last_set_dt),
                    "Account Expiration Date": format_datetime(acc_expires_dt),
                    "Account Expiration (days)": acc_expires_days if acc_expires_days is not None else "",
                    "Mobile": get_attr(entry, 'mobile', ''),
                    "Email": get_attr(entry, 'mail', ''),
                    "HomeDirectory": get_attr(entry, 'homeDirectory', ''),
                    "ProfilePath": get_attr(entry, 'profilePath', ''),
                    "ScriptPath": get_attr(entry, 'scriptPath', ''),
                    "UserAccountControl": uac,
                    "First Name": get_attr(entry, 'givenName', ''),
                    "Middle Name": get_attr(entry, 'middleName', ''),
                    "Last Name": get_attr(entry, 'sn', ''),
                    "Country": get_attr(entry, 'c', ''),
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                    "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                    "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                    "CanonicalName": get_attr(entry, 'canonicalName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting users: {e}")
            import traceback
            traceback.print_exc()

        self.results['Users'] = results
        logger.info(f"    Found {len(results)} users")
        return results

    def collect_user_spns(self) -> List[Dict]:
        """Collect user SPNs (Service Principal Names)."""
        logger.info("[-] Collecting User SPNs...")
        results = []

        try:
            # Users with SPNs - using different filter structure
            filter_str = "(&(objectCategory=person)(objectClass=user)(servicePrincipalName=*))"
            if self.config.only_enabled:
                filter_str = "(&(objectCategory=person)(objectClass=user)(servicePrincipalName=*)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"

            entries = self.search(
                self.base_dn,
                filter_str,
                ['sAMAccountName', 'name', 'description', 'memberOf',
                 'servicePrincipalName', 'primaryGroupID', 'pwdLastSet', 'userAccountControl']
            )

            for entry in entries:
                spns = get_attr_list(entry, 'servicePrincipalName')
                uac = get_attr(entry, 'userAccountControl', 0)
                uac_parsed = parse_uac(uac)

                pwd_last_set = get_attr(entry, 'pwdLastSet')
                # ldap3 returns datetime objects for these attributes
                pwd_last_set_dt = None
                if isinstance(pwd_last_set, datetime):
                    # Convert timezone-aware datetime to naive for comparison
                    pwd_last_set_dt = pwd_last_set.replace(tzinfo=None) if pwd_last_set.tzinfo else pwd_last_set

                # Get group memberships
                member_of = get_attr_list(entry, 'memberOf')
                groups = []
                for m in member_of:
                    match = re.search(r'CN=([^,]+)', str(m))
                    if match:
                        groups.append(match.group(1))

                for spn in spns:
                    spn_parts = str(spn).split('/')
                    service = spn_parts[0] if len(spn_parts) > 0 else ""
                    host = spn_parts[1] if len(spn_parts) > 1 else ""

                    results.append({
                        "Username": get_attr(entry, 'sAMAccountName', ''),
                        "Name": get_attr(entry, 'name', ''),
                        "Enabled": uac_parsed.get('Enabled', ''),
                        "Service": service,
                        "Host": host,
                        "Password Last Set": format_datetime(pwd_last_set_dt),
                        "Description": get_attr(entry, 'description', ''),
                        "Primary GroupID": get_attr(entry, 'primaryGroupID', ''),
                        "Memberof": ", ".join(groups),
                    })

        except Exception as e:
            logger.warning(f"Error collecting user SPNs: {e}")

        self.results['UserSPNs'] = results
        logger.info(f"    Found {len(results)} user SPNs")
        return results

    def collect_groups(self) -> List[Dict]:
        """Collect group objects."""
        logger.info("[-] Collecting Groups...")
        results = []

        try:
            entries = self.search(
                self.base_dn,
                "(objectCategory=group)",
                ['sAMAccountName', 'name', 'distinguishedName', 'canonicalName',
                 'description', 'groupType', 'adminCount', 'managedBy',
                 'objectSid', 'sIDHistory', 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                group_type = get_attr(entry, 'groupType')
                gt = safe_int(group_type)
                group_type_str = GROUP_TYPE.get(gt, "Unknown")
                if gt & 0x80000000:  # Security group
                    group_category = "Security"
                else:
                    group_category = "Distribution"

                if gt & 0x02:
                    group_scope = "Global"
                elif gt & 0x04:
                    group_scope = "DomainLocal"
                elif gt & 0x08:
                    group_scope = "Universal"
                else:
                    group_scope = "Unknown"

                # Managed by
                managed_by = get_attr(entry, 'managedBy', '')
                if managed_by:
                    match = re.search(r'CN=([^,]+)', str(managed_by))
                    if match:
                        managed_by = match.group(1)

                # SID History
                sid_history = get_attr_list(entry, 'sIDHistory')
                sid_history_str = ", ".join([sid_to_string(s) for s in sid_history]) if sid_history else ""

                results.append({
                    "Name": get_attr(entry, 'name', ''),
                    "SamAccountName": get_attr(entry, 'sAMAccountName', ''),
                    "Category": group_category,
                    "Scope": group_scope,
                    "AdminCount": get_attr(entry, 'adminCount', ''),
                    "Description": get_attr(entry, 'description', ''),
                    "ManagedBy": managed_by,
                    "SID": sid_to_string(get_attr(entry, 'objectSid')),
                    "SIDHistory": sid_history_str,
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                    "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                    "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                    "CanonicalName": get_attr(entry, 'canonicalName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting groups: {e}")

        self.results['Groups'] = results
        logger.info(f"    Found {len(results)} groups")
        return results

    def collect_group_members(self) -> List[Dict]:
        """Collect group memberships."""
        logger.info("[-] Collecting Group Members - May take some time...")
        results = []

        try:
            # Get all groups with members
            entries = self.search(
                self.base_dn,
                "(&(objectCategory=group)(member=*))",
                ['sAMAccountName', 'name', 'member', 'objectSid']
            )

            for entry in entries:
                group_name = get_attr(entry, 'name', '')
                group_sam = get_attr(entry, 'sAMAccountName', '')
                members = get_attr_list(entry, 'member')

                for member_dn in members:
                    # Extract member name from DN
                    match = re.search(r'CN=([^,]+)', str(member_dn))
                    member_name = match.group(1) if match else str(member_dn)

                    results.append({
                        "Group Name": group_name,
                        "Group SamAccountName": group_sam,
                        "Member Name": member_name,
                        "Member DN": str(member_dn),
                    })

        except Exception as e:
            logger.warning(f"Error collecting group members: {e}")

        self.results['GroupMembers'] = results
        logger.info(f"    Found {len(results)} group memberships")
        return results

    def collect_ous(self) -> List[Dict]:
        """Collect organizational units."""
        logger.info("[-] Collecting Organizational Units...")
        results = []

        try:
            entries = self.search(
                self.base_dn,
                "(objectCategory=organizationalUnit)",
                ['name', 'distinguishedName', 'description', 'gPLink',
                 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                gp_link = get_attr(entry, 'gPLink', '')

                results.append({
                    "Name": get_attr(entry, 'name', ''),
                    "Description": get_attr(entry, 'description', ''),
                    "gPLink": gp_link,
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                    "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                    "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting OUs: {e}")

        self.results['OUs'] = results
        logger.info(f"    Found {len(results)} organizational units")
        return results

    def collect_gpos(self) -> List[Dict]:
        """Collect Group Policy Objects."""
        logger.info("[-] Collecting GPOs...")
        results = []

        try:
            entries = self.search(
                self.base_dn,
                "(objectCategory=groupPolicyContainer)",
                ['displayName', 'name', 'distinguishedName', 'gPCFileSysPath',
                 'whenCreated', 'whenChanged', 'flags']
            )

            for entry in entries:
                flags = get_attr(entry, 'flags', 0)
                flags = safe_int(flags)

                # GPO flags
                user_disabled = bool(flags & 1)
                computer_disabled = bool(flags & 2)

                results.append({
                    "DisplayName": get_attr(entry, 'displayName', ''),
                    "GUID": get_attr(entry, 'name', ''),
                    "User Settings Disabled": user_disabled,
                    "Computer Settings Disabled": computer_disabled,
                    "GPCFileSysPath": get_attr(entry, 'gPCFileSysPath', ''),
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                    "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                    "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting GPOs: {e}")

        self.results['GPOs'] = results
        logger.info(f"    Found {len(results)} GPOs")
        return results

    def collect_gplinks(self) -> List[Dict]:
        """Collect GPO links (Scope of Management)."""
        logger.info("[-] Collecting gPLinks...")
        results = []

        try:
            # Get GPO dictionary for name lookup
            gpo_dict = {}
            gpo_entries = self.search(
                self.base_dn,
                "(objectCategory=groupPolicyContainer)",
                ['name', 'displayName']
            )
            for gpo in gpo_entries:
                gpo_name = safe_str(get_attr(gpo, 'name', '')).lower()
                gpo_display = safe_str(get_attr(gpo, 'displayName', ''))
                if gpo_name:
                    gpo_dict[gpo_name] = gpo_display

            # Get domain and OUs with gPLink
            entries = self.search(
                self.base_dn,
                "(|(objectCategory=domainDNS)(objectCategory=organizationalUnit))",
                ['name', 'distinguishedName', 'gPLink', 'gPOptions']
            )

            # Also get sites
            site_entries = self.search(
                f"CN=Sites,{self.config_dn}",
                "(objectCategory=site)",
                ['name', 'distinguishedName', 'gPLink', 'gPOptions']
            )
            entries.extend(site_entries)

            for entry in entries:
                gp_link = get_attr(entry, 'gPLink', '')
                if not gp_link:
                    continue

                gp_options = get_attr(entry, 'gPOptions', 0)
                block_inheritance = bool(safe_int(gp_options) & 1)

                # Parse gPLink - format: [LDAP://cn={GUID},cn=policies,...;options][...]
                links = re.findall(r'\[LDAP://([^;]+);(\d+)\]', str(gp_link), re.IGNORECASE)

                for link_dn, link_options in links:
                    link_options = int(link_options)
                    enforced = bool(link_options & 2)
                    disabled = bool(link_options & 1)

                    # Extract GPO GUID
                    guid_match = re.search(r'cn=(\{[^}]+\})', link_dn, re.IGNORECASE)
                    gpo_guid = guid_match.group(1) if guid_match else ""
                    gpo_display_name = gpo_dict.get(gpo_guid.lower(), gpo_guid)

                    results.append({
                        "SOM Name": get_attr(entry, 'name', ''),
                        "SOM DN": get_attr(entry, 'distinguishedName', ''),
                        "Block Inheritance": block_inheritance,
                        "GPO Name": gpo_display_name,
                        "GPO GUID": gpo_guid,
                        "Link Enabled": not disabled,
                        "Enforced": enforced,
                    })

        except Exception as e:
            logger.warning(f"Error collecting gPLinks: {e}")

        self.results['gPLinks'] = results
        logger.info(f"    Found {len(results)} GPO links")
        return results

    def collect_dns_zones(self) -> List[Dict]:
        """Collect DNS zones."""
        logger.info("[-] Collecting DNS Zones...")
        results = []

        try:
            # Check multiple locations for DNS zones
            search_bases = [
                self.base_dn,
                f"DC=DomainDnsZones,{self.base_dn}",
                f"DC=ForestDnsZones,{self.base_dn}",
            ]

            for search_base in search_bases:
                try:
                    entries = self.search(
                        search_base,
                        "(objectCategory=dnsZone)",
                        ['name', 'distinguishedName', 'whenCreated', 'whenChanged']
                    )

                    for entry in entries:
                        results.append({
                            "Name": get_attr(entry, 'name', ''),
                            "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                            "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                            "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                        })
                except Exception:
                    pass

        except Exception as e:
            logger.warning(f"Error collecting DNS zones: {e}")

        self.results['DNSZones'] = results
        logger.info(f"    Found {len(results)} DNS zones")
        return results

    def collect_dns_records(self) -> List[Dict]:
        """Collect DNS records."""
        logger.info("[-] Collecting DNS Records - May take some time...")
        results = []

        try:
            # Get DNS records from each zone
            if 'DNSZones' not in self.results:
                self.collect_dns_zones()

            for zone in self.results.get('DNSZones', []):
                zone_dn = zone.get('DistinguishedName', '')
                if not zone_dn:
                    continue

                try:
                    entries = self.search(
                        zone_dn,
                        "(objectCategory=dnsNode)",
                        ['name', 'dnsRecord', 'dNSTombstoned', 'whenCreated', 'whenChanged']
                    )

                    for entry in entries:
                        results.append({
                            "Zone": zone.get('Name', ''),
                            "Name": get_attr(entry, 'name', ''),
                            "Tombstoned": str(get_attr(entry, 'dNSTombstoned', '')),
                            "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                            "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                        })
                except Exception:
                    pass

        except Exception as e:
            logger.warning(f"Error collecting DNS records: {e}")

        self.results['DNSRecords'] = results
        logger.info(f"    Found {len(results)} DNS records")
        return results

    def collect_printers(self) -> List[Dict]:
        """Collect printer objects."""
        logger.info("[-] Collecting Printers...")
        results = []

        try:
            entries = self.search(
                self.base_dn,
                "(objectCategory=printQueue)",
                ['printerName', 'name', 'serverName', 'portName', 'driverName',
                 'driverVersion', 'printShareName', 'uNCName', 'url',
                 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                results.append({
                    "Name": get_attr(entry, 'name', ''),
                    "Server Name": get_attr(entry, 'serverName', ''),
                    "Share Name": get_attr(entry, 'printShareName', ''),
                    "Port Name": get_attr(entry, 'portName', ''),
                    "Driver Name": get_attr(entry, 'driverName', ''),
                    "Driver Version": get_attr(entry, 'driverVersion', ''),
                    "UNC Name": get_attr(entry, 'uNCName', ''),
                    "URL": get_attr(entry, 'url', ''),
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                    "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                })

        except Exception as e:
            logger.warning(f"Error collecting printers: {e}")

        self.results['Printers'] = results
        logger.info(f"    Found {len(results)} printers")
        return results

    def collect_computers(self) -> List[Dict]:
        """Collect computer objects."""
        logger.info("[-] Collecting Computers - May take some time...")
        results = []
        now = datetime.now()

        try:
            # Using objectCategory=computer instead of samAccountType
            filter_str = "(objectCategory=computer)"
            if self.config.only_enabled:
                filter_str = "(&(objectCategory=computer)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"

            entries = self.search(
                self.base_dn,
                filter_str,
                ['sAMAccountName', 'name', 'distinguishedName', 'dNSHostName',
                 'operatingSystem', 'operatingSystemVersion', 'operatingSystemServicePack',
                 'operatingSystemHotfix', 'userAccountControl', 'pwdLastSet',
                 'lastLogonTimestamp', 'description', 'primaryGroupID', 'objectSid',
                 'sIDHistory', 'servicePrincipalName', 'msDS-AllowedToDelegateTo',
                 'ms-ds-CreatorSid', 'msDS-SupportedEncryptionTypes',
                 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                uac = get_attr(entry, 'userAccountControl', 0)
                uac_parsed = parse_uac(uac)

                # Password last set
                pwd_last_set = get_attr(entry, 'pwdLastSet')
                # ldap3 returns datetime objects for these attributes
                pwd_last_set_dt = None
                if isinstance(pwd_last_set, datetime):
                    # Convert timezone-aware datetime to naive for comparison
                    pwd_last_set_dt = pwd_last_set.replace(tzinfo=None) if pwd_last_set.tzinfo else pwd_last_set
                    # Check if pwdLastSet is 0 (1601-01-01 epoch) which means password never set
                    if pwd_last_set_dt.year == 1601:
                        pwd_last_set_dt = None
                pwd_age_days = None
                pwd_not_changed_max = False

                if pwd_last_set_dt:
                    pwd_age_days = (now - pwd_last_set_dt).days
                    if pwd_age_days > self.config.password_age_days:
                        pwd_not_changed_max = True

                # Last logon
                last_logon = get_attr(entry, 'lastLogonTimestamp')
                # ldap3 returns datetime objects for these attributes
                last_logon_dt = None
                if isinstance(last_logon, datetime):
                    # Convert timezone-aware datetime to naive for comparison
                    last_logon_dt = last_logon.replace(tzinfo=None) if last_logon.tzinfo else last_logon
                    # Check if lastLogonTimestamp is 0 (1601-01-01 epoch) which means never logged in
                    if last_logon_dt.year == 1601:
                        last_logon_dt = None
                logon_age_days = None
                dormant = False

                if last_logon_dt:
                    logon_age_days = (now - last_logon_dt).days
                    if logon_age_days > self.config.dormant_days:
                        dormant = True
                    if logon_age_days > self.config.dormant_days:
                        dormant = True

                # Operating system
                os_name = safe_str(get_attr(entry, 'operatingSystem', ''))
                os_version = safe_str(get_attr(entry, 'operatingSystemVersion', ''))
                os_sp = safe_str(get_attr(entry, 'operatingSystemServicePack', ''))
                os_full = f"{os_name} {os_sp} {os_version}".strip()

                # Delegation
                delegation_type = None
                delegation_protocol = None
                delegation_services = None

                if uac_parsed.get('TrustedForDelegation', False):
                    delegation_type = "Unconstrained"
                    delegation_services = "Any"

                allowed_to_delegate = get_attr_list(entry, 'msDS-AllowedToDelegateTo')
                if allowed_to_delegate:
                    delegation_type = "Constrained"
                    delegation_services = ", ".join([str(s) for s in allowed_to_delegate])

                if uac_parsed.get('TrustedToAuthForDelegation', False):
                    delegation_protocol = "Any"
                elif delegation_type:
                    delegation_protocol = "Kerberos"

                # Kerberos encryption types
                enc_types = get_attr(entry, 'msDS-SupportedEncryptionTypes')
                kerb_enc = parse_kerb_enc_types(enc_types)

                # SID History
                sid_history = get_attr_list(entry, 'sIDHistory')
                sid_history_str = ", ".join([sid_to_string(s) for s in sid_history]) if sid_history else ""

                # Creator SID (for machine accounts created by users)
                creator_sid = get_attr(entry, 'ms-ds-CreatorSid')
                creator_sid_str = sid_to_string(creator_sid) if creator_sid else ""

                results.append({
                    "Name": get_attr(entry, 'name', ''),
                    "SamAccountName": get_attr(entry, 'sAMAccountName', ''),
                    "DNSHostName": get_attr(entry, 'dNSHostName', ''),
                    "Enabled": uac_parsed.get('Enabled', ''),
                    "Operating System": os_full,
                    "Description": get_attr(entry, 'description', ''),
                    "Logon Age (days)": logon_age_days,
                    "Password Age (days)": pwd_age_days,
                    f"Dormant (> {self.config.dormant_days} days)": dormant,
                    f"Password Age (> {self.config.password_age_days} days)": pwd_not_changed_max,
                    "Delegation Type": delegation_type or "",
                    "Delegation Protocol": delegation_protocol or "",
                    "Delegation Services": delegation_services or "",
                    "Kerberos RC4": kerb_enc.get('RC4', '') if kerb_enc.get('RC4') else '',
                    "Kerberos AES-128bit": kerb_enc.get('AES128', '') if kerb_enc.get('AES128') else '',
                    "Kerberos AES-256bit": kerb_enc.get('AES256', '') if kerb_enc.get('AES256') else '',
                    "Primary GroupID": get_attr(entry, 'primaryGroupID', ''),
                    "SID": sid_to_string(get_attr(entry, 'objectSid')),
                    "SIDHistory": sid_history_str,
                    "Creator SID": creator_sid_str,
                    "Last Logon Date": format_datetime(last_logon_dt),
                    "Password LastSet": format_datetime(pwd_last_set_dt),
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                    "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                    "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting computers: {e}")

        self.results['Computers'] = results
        logger.info(f"    Found {len(results)} computers")
        return results

    def collect_computer_spns(self) -> List[Dict]:
        """Collect computer SPNs."""
        logger.info("[-] Collecting Computer SPNs...")
        results = []

        try:
            filter_str = "(&(objectCategory=computer)(servicePrincipalName=*))"
            if self.config.only_enabled:
                filter_str = "(&(objectCategory=computer)(servicePrincipalName=*)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"

            entries = self.search(
                self.base_dn,
                filter_str,
                ['name', 'servicePrincipalName']
            )

            for entry in entries:
                spns = get_attr_list(entry, 'servicePrincipalName')
                computer_name = get_attr(entry, 'name', '')

                for spn in spns:
                    spn_parts = str(spn).split('/')
                    service = spn_parts[0] if len(spn_parts) > 0 else ""
                    host = spn_parts[1] if len(spn_parts) > 1 else ""

                    results.append({
                        "Computer": computer_name,
                        "Service": service,
                        "Host": host,
                        "SPN": str(spn),
                    })

        except Exception as e:
            logger.warning(f"Error collecting computer SPNs: {e}")

        self.results['ComputerSPNs'] = results
        logger.info(f"    Found {len(results)} computer SPNs")
        return results

    def collect_laps(self) -> List[Dict]:
        """Collect LAPS (Local Administrator Password Solution) information."""
        logger.info("[-] Collecting LAPS - Needs Privileged Account...")
        results = []

        try:
            # Check if LAPS is installed
            laps_entries = self.search(
                self.schema_dn,
                "(name=ms-Mcs-AdmPwd)",
                ['name']
            )

            if not laps_entries:
                logger.warning("[*] LAPS is not installed in this environment")
                return results

            # Get LAPS passwords
            entries = self.search(
                self.base_dn,
                "(objectCategory=computer)",
                ['name', 'dNSHostName', 'ms-Mcs-AdmPwd', 'ms-Mcs-AdmPwdExpirationTime',
                 'userAccountControl']
            )

            for entry in entries:
                uac = get_attr(entry, 'userAccountControl', 0)
                uac_parsed = parse_uac(uac)

                laps_pwd = safe_str(get_attr(entry, 'ms-Mcs-AdmPwd', ''))
                laps_exp = get_attr(entry, 'ms-Mcs-AdmPwdExpirationTime')
                laps_exp_int = safe_int(laps_exp)
                laps_exp_dt = windows_timestamp_to_datetime(laps_exp_int) if laps_exp_int else None

                # Only include if we can read the password or it's set
                password_readable = bool(laps_pwd)

                results.append({
                    "Computer Name": get_attr(entry, 'name', ''),
                    "DNSHostName": get_attr(entry, 'dNSHostName', ''),
                    "Enabled": uac_parsed.get('Enabled', ''),
                    "LAPS Password": laps_pwd if laps_pwd else "(Not Readable)",
                    "Password Expiration": str(laps_exp_dt) if laps_exp_dt else "",
                    "Password Readable": password_readable,
                })

        except Exception as e:
            logger.warning(f"Error collecting LAPS: {e}")

        self.results['LAPS'] = results
        logger.info(f"    Found {len(results)} LAPS entries")
        return results

    def collect_bitlocker(self) -> List[Dict]:
        """Collect BitLocker recovery keys."""
        logger.info("[-] Collecting BitLocker Recovery Keys - Needs Privileged Account...")
        results = []

        try:
            entries = self.search(
                self.base_dn,
                "(objectCategory=msFVE-RecoveryInformation)",
                ['distinguishedName', 'name', 'msFVE-RecoveryPassword',
                 'msFVE-RecoveryGuid', 'msFVE-VolumeGuid', 'whenCreated']
            )

            for entry in entries:
                dn = get_attr(entry, 'distinguishedName', '')
                # Extract computer name from DN
                computer_match = re.search(r'CN=([^,]+),CN=([^,]+)', str(dn))
                computer_name = computer_match.group(2) if computer_match else ""

                results.append({
                    "Computer Name": computer_name,
                    "Name": get_attr(entry, 'name', ''),
                    "Recovery Password": get_attr(entry, 'msFVE-RecoveryPassword', ''),
                    "Recovery GUID": get_attr(entry, 'msFVE-RecoveryGuid', ''),
                    "Volume GUID": get_attr(entry, 'msFVE-VolumeGuid', ''),
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                })

        except Exception as e:
            logger.warning(f"Error collecting BitLocker keys: {e}")

        self.results['BitLocker'] = results
        logger.info(f"    Found {len(results)} BitLocker recovery keys")
        return results

    def collect_kerberoast(self) -> List[Dict]:
        """Collect Kerberoastable accounts and request TGS tickets."""
        logger.info("[-] Collecting Kerberoast - Requesting TGS tickets...")
        results = []

        if not IMPACKET_AVAILABLE:
            logger.warning("[*] impacket not available - Kerberoast disabled")
            return results

        try:
            # Get users with SPNs (excluding computers)
            filter_str = "(&(objectCategory=person)(objectClass=user)(servicePrincipalName=*)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"

            entries = self.search(
                self.base_dn,
                filter_str,
                ['sAMAccountName', 'servicePrincipalName', 'distinguishedName']
            )

            for entry in entries:
                spns = get_attr_list(entry, 'servicePrincipalName')
                username = get_attr(entry, 'sAMAccountName', '')
                user_domain = dn_to_fqdn(str(get_attr(entry, 'distinguishedName', '')))

                for spn in spns:
                    # Note: Getting actual TGS tickets requires domain membership or valid TGT
                    # This is a placeholder - actual implementation would use impacket's GetUserSPNs
                    results.append({
                        "Username": username,
                        "ServicePrincipalName": str(spn),
                        "Domain": user_domain,
                        "John": f"$krb5tgs${spn}:<hash_would_be_here>",
                        "Hashcat": f"$krb5tgs$23$*{username}${user_domain}${spn}*$<hash>",
                    })

        except Exception as e:
            logger.warning(f"Error collecting Kerberoast: {e}")

        self.results['Kerberoast'] = results
        logger.info(f"    Found {len(results)} Kerberoastable accounts")
        return results

    def collect_schema_history(self) -> List[Dict]:
        """Collect schema history/updates."""
        logger.info("[-] Collecting Schema History - May take some time...")
        results = []

        try:
            entries = self.search(
                self.schema_dn,
                "(objectCategory=*)",
                ['name', 'objectClass', 'whenCreated', 'whenChanged'],
            )

            for entry in entries:
                obj_class = get_attr(entry, 'objectClass')
                if isinstance(obj_class, list):
                    obj_class = ", ".join([str(c) for c in obj_class])

                results.append({
                    "ObjectClass": obj_class,
                    "Name": get_attr(entry, 'name', ''),
                    "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                    "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                    "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting schema history: {e}")

        self.results['SchemaHistory'] = results
        logger.info(f"    Found {len(results)} schema objects")
        return results

    def run(self):
        """Run the AD reconnaissance."""
        print(BANNER)
        logger.info(f"Starting PyADRecon at {self.start_time}")
        logger.info(f"Target: {self.config.domain_controller}")
        logger.info(f"Authentication: {self.config.auth_method.upper()}")

        if not self.connect():
            logger.error("Failed to connect to domain controller")
            return False

        logger.info(f"[*] Commencing - {datetime.now()}")

        # Collect data based on configuration
        if self.config.collect_domain:
            self.collect_domain_info()

        if self.config.collect_forest:
            self.collect_forest_info()

        if self.config.collect_trusts:
            self.collect_trusts()

        if self.config.collect_sites:
            self.collect_sites()

        if self.config.collect_subnets:
            self.collect_subnets()

        if self.config.collect_schema:
            self.collect_schema_history()

        if self.config.collect_password_policy:
            self.collect_password_policy()

        if self.config.collect_fgpp:
            self.collect_fine_grained_password_policies()

        if self.config.collect_dcs:
            self.collect_domain_controllers()

        if self.config.collect_users:
            self.collect_users()

        if self.config.collect_user_spns:
            self.collect_user_spns()

        if self.config.collect_groups:
            self.collect_groups()

        if self.config.collect_group_members:
            self.collect_group_members()

        if self.config.collect_ous:
            self.collect_ous()

        if self.config.collect_gpos:
            self.collect_gpos()

        if self.config.collect_gplinks:
            self.collect_gplinks()

        if self.config.collect_dns_zones:
            self.collect_dns_zones()

        if self.config.collect_dns_records:
            self.collect_dns_records()

        if self.config.collect_printers:
            self.collect_printers()

        if self.config.collect_computers:
            self.collect_computers()

        if self.config.collect_computer_spns:
            self.collect_computer_spns()

        if self.config.collect_laps:
            self.collect_laps()

        if self.config.collect_bitlocker:
            self.collect_bitlocker()

        if self.config.collect_kerberoast:
            self.collect_kerberoast()

        # Calculate execution time
        end_time = datetime.now()
        duration = end_time - self.start_time
        logger.info(f"[*] Total Execution Time: {duration}")

        return True

    def _sanitize_value(self, value):
        """Sanitize a value for export, handling ldap3 Attribute objects."""
        # Handle ldap3 Attribute objects
        if hasattr(value, 'raw_values'):
            value = value.value
        if value is None:
            return ''
        elif isinstance(value, bytes):
            return value.decode('utf-8', errors='replace')
        elif isinstance(value, (list, dict)):
            return str(value)
        elif isinstance(value, (str, int, float, bool)):
            return value
        else:
            return str(value)

    def _sanitize_results(self):
        """Sanitize all results for export."""
        for name, data in self.results.items():
            for row in data:
                for key in row:
                    row[key] = self._sanitize_value(row[key])

    def export_csv(self, output_dir: str):
        """Export results to CSV files."""
        logger.info("[*] Exporting to CSV...")

        # Sanitize all results before export
        self._sanitize_results()

        csv_dir = os.path.join(output_dir, "CSV-Files")
        os.makedirs(csv_dir, exist_ok=True)

        for name, data in self.results.items():
            if not data:
                continue

            filename = os.path.join(csv_dir, f"{name}.csv")
            try:
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    if data:
                        writer = csv.DictWriter(f, fieldnames=data[0].keys())
                        writer.writeheader()
                        writer.writerows(data)
                logger.info(f"    Exported {name}.csv ({len(data)} records)")
            except Exception as e:
                logger.warning(f"    Failed to export {name}.csv: {e}")

        return csv_dir

    def export_xlsx(self, output_dir: str, domain_name: str = ""):
        """Export results to Excel file (optimized for large datasets)."""
        if not OPENPYXL_AVAILABLE:
            logger.warning("[*] openpyxl not available - Excel export disabled")
            return None

        logger.info("[*] Generating Excel Report (optimized mode)...")
        start_time = datetime.now()

        try:
            # Use write_only mode for much faster performance
            from openpyxl import Workbook
            from openpyxl.cell import WriteOnlyCell

            wb = Workbook(write_only=True)

            # Define styles for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

            # Create Table of Contents sheet first (small, use regular mode)
            toc_data = [
                ["PyADRecon Report"],
                [f"Generated: {datetime.now()}"],
                [f"Domain: {domain_name or dn_to_fqdn(self.base_dn)}"],
                [""],
                ["Sheet Name", "Record Count"],
            ]
            for name in self.results.keys():
                if self.results[name]:
                    toc_data.append([name, len(self.results[name])])

            toc_ws = wb.create_sheet("Table of Contents")
            for row in toc_data:
                toc_ws.append(row)

            # Process each result set
            total_sheets = len([k for k, v in self.results.items() if v])
            current_sheet = 0

            for name, data in self.results.items():
                if not data:
                    continue

                current_sheet += 1
                record_count = len(data)
                logger.info(f"    [{current_sheet}/{total_sheets}] Writing {name} ({record_count:,} records)...")

                ws = wb.create_sheet(name[:31])  # Excel sheet name limit

                headers = list(data[0].keys())

                # Write header row with styling
                header_row = []
                for header in headers:
                    cell = WriteOnlyCell(ws, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    header_row.append(cell)
                ws.append(header_row)

                # Write data rows in batches for progress reporting
                batch_size = 10000
                for i, row_data in enumerate(data):
                    row = []
                    for header in headers:
                        value = row_data.get(header, '')
                        # Handle ldap3 Attribute objects
                        if hasattr(value, 'raw_values'):
                            value = value.value
                        # Convert to safe types
                        if value is None:
                            value = ''
                        elif isinstance(value, bytes):
                            value = value.decode('utf-8', errors='replace')
                        elif isinstance(value, (list, dict)):
                            value = str(value)
                        elif not isinstance(value, (str, int, float, bool)):
                            value = str(value)
                        row.append(value)
                    ws.append(row)

                    # Progress for large datasets
                    if record_count > batch_size and (i + 1) % batch_size == 0:
                        pct = ((i + 1) / record_count) * 100
                        logger.info(f"        {i + 1:,}/{record_count:,} rows ({pct:.0f}%)")

            # Save workbook
            filename = os.path.join(output_dir, f"{domain_name or 'ADRecon'}-Report.xlsx")
            logger.info(f"    Saving Excel file...")
            wb.save(filename)

            duration = datetime.now() - start_time
            logger.info(f"[+] Excel Report saved to: {filename} (took {duration})")
            return filename

        except Exception as e:
            logger.error(f"Failed to create Excel report: {e}")
            import traceback
            traceback.print_exc()
            return None

    def close(self):
        """Close LDAP connection."""
        if self.conn:
            self.conn.unbind()            


def generate_excel_from_csv(csv_dir: str, output_file: str = None):
    """
    Standalone function to generate Excel report from CSV files.
    This is optimized for large datasets and can be run independently.

    Args:
        csv_dir: Path to directory containing CSV files
        output_file: Output Excel file path (optional, defaults to same directory)
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("[!] openpyxl not available - install with: pip install openpyxl")
        return None

    if not os.path.isdir(csv_dir):
        logger.error(f"[!] CSV directory not found: {csv_dir}")
        return None

    logger.info(f"[*] Generating Excel Report from CSV files in: {csv_dir}")
    start_time = datetime.now()

    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell

        # Find all CSV files
        csv_files = sorted([f for f in os.listdir(csv_dir) if f.endswith('.csv')])
        if not csv_files:
            logger.error(f"[!] No CSV files found in: {csv_dir}")
            return None

        logger.info(f"    Found {len(csv_files)} CSV files")

        # Use write_only mode for performance
        wb = Workbook(write_only=True)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

        # Build TOC data
        toc_data = [
            ["PyADRecon Report"],
            [f"Generated: {datetime.now()}"],
            [f"Source: {csv_dir}"],
            [""],
            ["Sheet Name", "Record Count"],
        ]

        # First pass - count records for TOC
        file_counts = {}
        for csv_file in csv_files:
            csv_path = os.path.join(csv_dir, csv_file)
            with open(csv_path, 'r', encoding='utf-8', errors='replace') as f:
                count = sum(1 for _ in f) - 1  # Subtract header
                file_counts[csv_file] = max(0, count)
            sheet_name = csv_file.replace('.csv', '')
            toc_data.append([sheet_name, file_counts[csv_file]])

        # Create TOC sheet
        toc_ws = wb.create_sheet("Table of Contents")
        for row in toc_data:
            toc_ws.append(row)

        # Process each CSV file
        total_files = len(csv_files)
        for idx, csv_file in enumerate(csv_files, 1):
            sheet_name = csv_file.replace('.csv', '')[:31]  # Excel limit
            record_count = file_counts[csv_file]

            logger.info(f"    [{idx}/{total_files}] Processing {csv_file} ({record_count:,} records)...")

            csv_path = os.path.join(csv_dir, csv_file)
            ws = wb.create_sheet(sheet_name)

            with open(csv_path, 'r', encoding='utf-8', errors='replace', newline='') as f:
                reader = csv.reader(f)

                # Write header with styling
                try:
                    headers = next(reader)
                    header_row = []
                    for header in headers:
                        cell = WriteOnlyCell(ws, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        header_row.append(cell)
                    ws.append(header_row)
                except StopIteration:
                    continue  # Empty file

                # Write data rows
                batch_size = 10000
                row_num = 0
                for row in reader:
                    # Convert empty strings and handle encoding
                    clean_row = []
                    for val in row:
                        if val == '':
                            clean_row.append('')
                        else:
                            # Try to convert to number if possible
                            try:
                                if '.' in val:
                                    clean_row.append(float(val))
                                else:
                                    clean_row.append(int(val))
                            except ValueError:
                                clean_row.append(val)
                    ws.append(clean_row)
                    row_num += 1

                    # Progress for large files
                    if record_count > batch_size and row_num % batch_size == 0:
                        pct = (row_num / record_count) * 100
                        logger.info(f"        {row_num:,}/{record_count:,} rows ({pct:.0f}%)")

        # Determine output filename
        if output_file:
            filename = output_file
        else:
            # Put in parent directory of CSV-Files
            parent_dir = os.path.dirname(csv_dir.rstrip('/'))
            if os.path.basename(csv_dir) == 'CSV-Files':
                filename = os.path.join(parent_dir, "ADRecon-Report.xlsx")
            else:
                filename = os.path.join(csv_dir, "ADRecon-Report.xlsx")

        logger.info(f"    Saving Excel file...")
        wb.save(filename)

        duration = datetime.now() - start_time
        logger.info(f"[+] Excel Report saved to: {filename}")
        logger.info(f"[+] Total time: {duration}")
        return filename

    except Exception as e:
        logger.error(f"Failed to create Excel report: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    parser = argparse.ArgumentParser(
        description="PyADRecon - Python Active Directory Reconnaissance Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage with NTLM authentication
  %(prog)s -dc 192.168.1.1 -u admin -p password123 -d DOMAIN.LOCAL

  # With Kerberos authentication
  %(prog)s -dc dc01.domain.local -u admin -p password123 -d DOMAIN.LOCAL --auth kerberos

  # Only collect specific modules
  %(prog)s -dc 192.168.1.1 -u admin -p pass -d DOMAIN.LOCAL --collect users,groups,computers

  # Output to specific directory
  %(prog)s -dc 192.168.1.1 -u admin -p pass -d DOMAIN.LOCAL -o /tmp/adrecon_output

  # Generate Excel report from existing CSV files (standalone mode)
  %(prog)s --generate-excel-from /path/to/CSV-Files -o report.xlsx
        """
    )

    # Standalone Excel generation mode
    parser.add_argument('--generate-excel-from', metavar='CSV_DIR',
                       help='Generate Excel report from CSV directory (standalone mode, no AD connection needed)')

    # Required arguments (not required if using --generate-excel-from)
    parser.add_argument('-dc', '--domain-controller', default='',
                       help='Domain Controller IP or hostname')
    parser.add_argument('-u', '--username', default='',
                       help='Username for authentication')
    parser.add_argument('-p', '--password', default='',
                       help='Password for authentication')

    # Optional arguments
    parser.add_argument('-d', '--domain', default='',
                       help='Domain name (e.g., DOMAIN.LOCAL)')
    parser.add_argument('--auth', choices=['ntlm', 'kerberos'], default='ntlm',
                       help='Authentication method (default: ntlm)')
    parser.add_argument('--ssl', action='store_true',
                       help='Use SSL/TLS (LDAPS)')
    parser.add_argument('--port', type=int, default=389,
                       help='LDAP port (default: 389, use 636 for LDAPS)')
    parser.add_argument('-o', '--output', default='',
                       help='Output directory (default: PyADRecon-Report-<timestamp>)')
    parser.add_argument('--page-size', type=int, default=500,
                       help='LDAP page size (default: 500)')
    parser.add_argument('--threads', type=int, default=10,
                       help='Number of threads (default: 10)')
    parser.add_argument('--dormant-days', type=int, default=90,
                       help='Days for dormant account threshold (default: 90)')
    parser.add_argument('--password-age', type=int, default=30,
                       help='Days for password age threshold (default: 30)')
    parser.add_argument('--only-enabled', action='store_true',
                       help='Only collect enabled objects')
    parser.add_argument('--collect', default='default',
                       help='Comma-separated modules to collect (default: all except kerberoast,acls)')
    parser.add_argument('--no-excel', action='store_true',
                       help='Skip Excel report generation')
    parser.add_argument('-v', '--verbose', action='store_true',
                       help='Verbose output')

    args = parser.parse_args()

    if args.verbose:
        logger.setLevel(logging.DEBUG)

    # Handle standalone Excel generation mode
    if args.generate_excel_from:
        print(BANNER)
        logger.info("Running in standalone Excel generation mode")

        if not OPENPYXL_AVAILABLE:
            print("[!] openpyxl library required: pip install openpyxl")
            sys.exit(1)

        csv_dir = args.generate_excel_from
        output_file = args.output if args.output else None

        result = generate_excel_from_csv(csv_dir, output_file)
        if result:
            sys.exit(0)
        else:
            sys.exit(1)

    # Check required arguments for normal mode
    if not args.domain_controller or not args.username or not args.password:
        print("[!] Error: -dc, -u, and -p are required for AD reconnaissance mode")
        print("[!] Use --generate-excel-from for standalone Excel generation from CSV files")
        sys.exit(1)

    # Check for required library
    if not LDAP3_AVAILABLE:
        print("[!] ldap3 library required: pip install ldap3")
        print("[!] Install all dependencies: pip install -r requirements.txt")
        sys.exit(1)

    # Parse collection modules
    collect_modules = args.collect.lower().split(',')

    config = ADReconConfig(
        domain_controller=args.domain_controller,
        domain=args.domain,
        username=args.username,
        password=args.password,
        auth_method=args.auth,
        use_ssl=args.ssl,
        port=636 if args.ssl else args.port,
        page_size=args.page_size,
        threads=args.threads,
        dormant_days=args.dormant_days,
        password_age_days=args.password_age,
        only_enabled=args.only_enabled,
    )

    # Configure collection based on modules
    if 'default' in collect_modules or 'all' in collect_modules:
        pass  # Use defaults
    else:
        # Disable all first
        config.collect_forest = 'forest' in collect_modules
        config.collect_domain = 'domain' in collect_modules
        config.collect_trusts = 'trusts' in collect_modules
        config.collect_sites = 'sites' in collect_modules
        config.collect_subnets = 'subnets' in collect_modules
        config.collect_schema = 'schema' in collect_modules or 'schemahistory' in collect_modules
        config.collect_password_policy = 'passwordpolicy' in collect_modules
        config.collect_fgpp = 'fgpp' in collect_modules or 'finegrainedpasswordpolicy' in collect_modules
        config.collect_dcs = 'dcs' in collect_modules or 'domaincontrollers' in collect_modules
        config.collect_users = 'users' in collect_modules
        config.collect_user_spns = 'userspns' in collect_modules
        config.collect_groups = 'groups' in collect_modules
        config.collect_group_members = 'groupmembers' in collect_modules
        config.collect_ous = 'ous' in collect_modules
        config.collect_gpos = 'gpos' in collect_modules
        config.collect_gplinks = 'gplinks' in collect_modules
        config.collect_dns_zones = 'dnszones' in collect_modules
        config.collect_dns_records = 'dnsrecords' in collect_modules
        config.collect_printers = 'printers' in collect_modules
        config.collect_computers = 'computers' in collect_modules
        config.collect_computer_spns = 'computerspns' in collect_modules
        config.collect_laps = 'laps' in collect_modules
        config.collect_bitlocker = 'bitlocker' in collect_modules
        config.collect_kerberoast = 'kerberoast' in collect_modules
        config.collect_acls = 'acls' in collect_modules

    # Create output directory
    if args.output:
        output_dir = args.output
    else:
        output_dir = f"PyADRecon-Report-{datetime.now().strftime('%Y%m%d%H%M%S')}"

    os.makedirs(output_dir, exist_ok=True)
    config.output_dir = output_dir

    # Run reconnaissance
    recon = PyADRecon(config)

    try:
        if recon.run():
            # Export results
            csv_dir = recon.export_csv(output_dir)

            if not args.no_excel and OPENPYXL_AVAILABLE:
                domain_name = config.domain or dn_to_fqdn(recon.base_dn)
                recon.export_xlsx(output_dir, domain_name.replace('.', '_'))

            logger.info(f"[*] Output Directory: {os.path.abspath(output_dir)}")
            logger.info("[*] Completed.")
        else:
            logger.error("[!] Reconnaissance failed")
            sys.exit(1)

    except KeyboardInterrupt:
        logger.warning("\n[!] Interrupted by user")
        sys.exit(1)
    finally:
        recon.close()


if __name__ == "__main__":
    main()
